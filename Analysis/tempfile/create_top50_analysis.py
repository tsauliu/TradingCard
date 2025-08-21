#!/usr/bin/env python3
"""
Top 50 Products Analysis from tcg_prices_bda
Creates 2-sheet Excel: Summary and Detailed data
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def add_formatted_data(ws, df, start_row, title, price_format=False):
    """Add formatted data to worksheet"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add data starting from next row
    data_start_row = start_row + 2
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Add data rows
    last_data_row = data_start_row
    for row_idx, row in enumerate(df.itertuples(index=False), data_start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, (int, float)) and col_idx > 1:
                if price_format:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        last_data_row = row_idx
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check header length
        header_value = ws.cell(row=data_start_row, column=col_idx).value
        if header_value:
            max_length = max(max_length, len(str(header_value)))
        
        # Check data lengths
        for check_row in range(data_start_row + 1, last_data_row + 1):
            cell_value = ws.cell(row=check_row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
    
    return last_data_row + 3

def main():
    # Set up credentials
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/Analysis/service-account.json'
    client = bigquery.Client()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    # Step 1: Get top 50 products by lifecycle quantity_sold with recent week avg price > $5
    top_products_query = """
    WITH recent_week_prices AS (
      SELECT 
        product_id,
        DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
        AVG(market_price) as avg_market_price
      FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda`
      WHERE market_price > 0 
        AND bucket_start_date IS NOT NULL
        AND bucket_start_date >= '2024-01-01'
        AND product_id IS NOT NULL
      GROUP BY product_id, DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    ),
    latest_week_per_product AS (
      SELECT 
        product_id,
        MAX(week_start) as latest_week
      FROM recent_week_prices
      GROUP BY product_id
    ),
    products_with_recent_price AS (
      SELECT 
        rwp.product_id,
        rwp.avg_market_price as recent_avg_price
      FROM recent_week_prices rwp
      INNER JOIN latest_week_per_product lwp 
        ON rwp.product_id = lwp.product_id 
        AND rwp.week_start = lwp.latest_week
      WHERE rwp.avg_market_price > 5
    )
    SELECT 
        p.product_id,
        m.product_name,
        SUM(CAST(p.total_quantity_sold AS INT64)) as lifecycle_quantity_sold,
        pwrp.recent_avg_price
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` p
    INNER JOIN products_with_recent_price pwrp 
      ON p.product_id = pwrp.product_id
    LEFT JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` m
      ON CAST(p.product_id AS STRING) = CAST(m.product_productId AS STRING)
    WHERE p.total_quantity_sold IS NOT NULL 
      AND p.total_quantity_sold != ''
      AND (m.product_name IS NULL OR (
        LOWER(m.product_name) NOT LIKE '%pack%'
        AND LOWER(m.product_name) NOT LIKE '%bundle%'
        AND LOWER(m.product_name) NOT LIKE '%box%'
        AND LOWER(m.product_name) NOT LIKE '%booster%'
        AND LOWER(m.product_name) NOT LIKE '%sealed%'
        AND LOWER(m.product_name) NOT LIKE '%deck%'
        AND LOWER(m.product_name) NOT LIKE '%tin%'
        AND LOWER(m.product_name) NOT LIKE '%collection%'
        AND LOWER(m.product_name) NOT LIKE '%set%'
        AND LOWER(m.product_name) NOT LIKE '%lot%'
        AND LOWER(m.product_name) NOT LIKE '%theme%'
        AND LOWER(m.product_name) NOT LIKE '%starter%'
      ))
    GROUP BY p.product_id, m.product_name, pwrp.recent_avg_price
    ORDER BY lifecycle_quantity_sold DESC
    LIMIT 50
    """
    
    print("Getting top 50 single cards by lifecycle quantity_sold (excluding packs/bundles)...")
    top_products_df = client.query(top_products_query).to_dataframe()
    
    if top_products_df.empty:
        print("No products found. Exiting.")
        return
    
    print(f"Found {len(top_products_df)} top products")
    
    # Get the list of top 50 product IDs
    top_product_ids = top_products_df['product_id'].tolist()
    top_product_ids_str = "', '".join([str(pid) for pid in top_product_ids])
    
    # Step 2: Get weekly data for these products with product names
    weekly_data_query = f"""
    SELECT 
        p.product_id,
        m.product_name,
        p.condition,
        DATE_TRUNC(p.bucket_start_date, WEEK(MONDAY)) as week_start,
        AVG(p.market_price) as avg_market_price,
        SUM(p.quantity_sold) as weekly_quantity_sold
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` p
    LEFT JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` m
      ON CAST(p.product_id AS STRING) = CAST(m.product_productId AS STRING)
    WHERE p.product_id IN ('{top_product_ids_str}')
      AND p.bucket_start_date IS NOT NULL
      AND p.bucket_start_date >= '2024-01-01'
      AND p.condition IS NOT NULL
      AND p.condition != ''
      AND p.market_price > 0
      AND p.quantity_sold IS NOT NULL
    GROUP BY p.product_id, m.product_name, p.condition, DATE_TRUNC(p.bucket_start_date, WEEK(MONDAY))
    ORDER BY p.product_id, week_start, p.condition
    """
    
    print("Getting weekly data for top 50 products...")
    weekly_data_df = client.query(weekly_data_query).to_dataframe()
    
    if weekly_data_df.empty:
        print("No weekly data found. Exiting.")
        return
    
    print(f"Retrieved {len(weekly_data_df)} weekly data records")
    
    # Step 3: Calculate weighted averages and totals by week and product
    print("Calculating weighted averages...")
    
    # Calculate weighted price and total quantity per product per week
    weekly_summary = []
    
    for product_id in top_product_ids:
        product_data = weekly_data_df[weekly_data_df['product_id'] == product_id]
        
        if product_data.empty:
            continue
        
        # Get product name (should be same for all rows of this product)
        product_name = product_data['product_name'].iloc[0] if not product_data['product_name'].isna().all() else f"Product_{product_id}"
            
        for week in product_data['week_start'].unique():
            week_data = product_data[product_data['week_start'] == week]
            
            # Calculate weighted average price
            total_value = (week_data['avg_market_price'] * week_data['weekly_quantity_sold']).sum()
            total_quantity = week_data['weekly_quantity_sold'].sum()
            
            if total_quantity > 0:
                weighted_price = total_value / total_quantity
            else:
                weighted_price = 0
            
            weekly_summary.append({
                'product_id': product_id,
                'product_name': product_name,
                'week_start': week,
                'weighted_avg_price': round(weighted_price, 2),
                'total_weekly_quantity': int(total_quantity)
            })
    
    weekly_summary_df = pd.DataFrame(weekly_summary)
    
    # Step 4: Create pivot tables for summary sheet
    print("Creating pivot tables...")
    
    # Create a product lookup for names
    product_names = weekly_summary_df.drop_duplicates(['product_id', 'product_name'])[['product_id', 'product_name']]
    
    # Pivot weighted prices - weeks as columns
    price_pivot = weekly_summary_df.pivot(index='product_id', columns='week_start', values='weighted_avg_price')
    price_pivot = price_pivot.round(2).fillna(0)
    price_pivot.columns = [f"{col.strftime('%Y-%m-%d')}" for col in price_pivot.columns]
    price_pivot = price_pivot.reset_index()
    
    # Add product names to price pivot
    price_pivot = price_pivot.merge(product_names, on='product_id', how='left')
    price_pivot = price_pivot[['product_id', 'product_name'] + [col for col in price_pivot.columns if col not in ['product_id', 'product_name']]]
    price_pivot.columns.name = None
    
    # Pivot quantities - weeks as columns
    quantity_pivot = weekly_summary_df.pivot(index='product_id', columns='week_start', values='total_weekly_quantity')
    quantity_pivot = quantity_pivot.fillna(0).astype(int)
    quantity_pivot.columns = [f"{col.strftime('%Y-%m-%d')}" for col in quantity_pivot.columns]
    quantity_pivot = quantity_pivot.reset_index()
    
    # Add product names to quantity pivot
    quantity_pivot = quantity_pivot.merge(product_names, on='product_id', how='left')
    quantity_pivot = quantity_pivot[['product_id', 'product_name'] + [col for col in quantity_pivot.columns if col not in ['product_id', 'product_name']]]
    quantity_pivot.columns.name = None
    
    # Create output directory if it doesn't exist
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create Excel file
    excel_file = f"{output_dir}/{timestamp}_top50_products_analysis.xlsx"
    
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    current_row = 1
    current_row = add_formatted_data(ws_summary, price_pivot, current_row, 
                                   "Top 50 Products - Weekly Weighted Average Price", price_format=True)
    current_row = add_formatted_data(ws_summary, quantity_pivot, current_row, 
                                   "Top 50 Products - Weekly Total Quantity Sold", price_format=False)
    
    # Detailed Sheet
    ws_detailed = wb.create_sheet(title="Detailed Data")
    
    # Format detailed data
    detailed_df = weekly_data_df.copy()
    # Convert week_start to string format if it's datetime
    if pd.api.types.is_datetime64_any_dtype(detailed_df['week_start']):
        detailed_df['week_start'] = detailed_df['week_start'].dt.strftime('%Y-%m-%d')
    detailed_df = detailed_df.round({'avg_market_price': 2})
    
    add_formatted_data(ws_detailed, detailed_df, 1, "Detailed Weekly Data - All Products by Condition")
    
    wb.save(excel_file)
    
    print(f"Excel file saved: {excel_file}")
    print(f"Summary - Price pivot shape: {price_pivot.shape}")
    print(f"Summary - Quantity pivot shape: {quantity_pivot.shape}")
    print(f"Detailed data shape: {detailed_df.shape}")
    
    print(f"\nTop 10 single cards by lifecycle quantity (recent avg price > $5, excluding packs/bundles):")
    print(top_products_df.head(10)[['product_id', 'product_name', 'lifecycle_quantity_sold', 'recent_avg_price']])

if __name__ == "__main__":
    main()