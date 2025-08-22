#!/usr/bin/env python3
"""
Create BDA weekly summary with ASP and trading volume by condition
Weekly data with weeks as columns, formatted Excel output
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def add_formatted_data(ws, df, start_row, title):
    """Add formatted data to worksheet with simple formatting"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    # Add data starting from next row
    data_start_row = start_row + 2
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Add data rows
    last_data_row = data_start_row
    for row_idx, row in enumerate(df.itertuples(index=False), data_start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, (int, float)) and col_idx > 1:
                if 'ASP' in title or 'Price' in title:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        last_data_row = row_idx
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check header length
        header_value = ws.cell(row=data_start_row, column=col_idx).value
        if header_value:
            max_length = max(max_length, len(str(header_value)))
        
        # Check data lengths
        for check_row in range(data_start_row + 1, last_data_row + 1):
            cell_value = ws.cell(row=check_row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
    
    return last_data_row + 3

def main():
    # Set up credentials
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/Analysis/service-account.json'
    client = bigquery.Client()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    # Query for weekly ASP by condition (pivot format)
    asp_query = """
    WITH weekly_data AS (
      SELECT 
        condition,
        DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
        AVG(market_price) as avg_market_price
      FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda`
      WHERE market_price > 0 
        AND condition IS NOT NULL 
        AND condition != ''
        AND bucket_start_date IS NOT NULL
        AND bucket_start_date >= '2024-01-01'
      GROUP BY condition, DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    )
    SELECT *
    FROM weekly_data
    ORDER BY condition, week_start
    """
    
    # Query for weekly volume by condition (pivot format)  
    volume_query = """
    WITH weekly_data AS (
      SELECT 
        condition,
        DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
        SUM(quantity_sold) as quantity_sold
      FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda`
      WHERE condition IS NOT NULL 
        AND condition != ''
        AND quantity_sold IS NOT NULL
        AND bucket_start_date IS NOT NULL
        AND bucket_start_date >= '2024-01-01'
      GROUP BY condition, DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    )
    SELECT *
    FROM weekly_data
    ORDER BY condition, week_start
    """
    
    # Query for weekly unique products with sales (pivot format)
    unique_products_query = """
    WITH weekly_data AS (
      SELECT 
        DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
        COUNT(DISTINCT product_id) as unique_products_with_sales
      FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda`
      WHERE quantity_sold > 0
        AND product_id IS NOT NULL
        AND bucket_start_date IS NOT NULL
        AND bucket_start_date >= '2024-01-01'
      GROUP BY DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    )
    SELECT 
      'Total' as metric,
      week_start,
      unique_products_with_sales
    FROM weekly_data
    ORDER BY week_start
    """
    
    print("Executing ASP query...")
    asp_df = client.query(asp_query).to_dataframe()
    
    print("Executing volume query...")
    volume_df = client.query(volume_query).to_dataframe()
    
    print("Executing unique products query...")
    unique_products_df = client.query(unique_products_query).to_dataframe()
    
    print(f"ASP data: {len(asp_df)} rows")
    print(f"Volume data: {len(volume_df)} rows")
    print(f"Unique products data: {len(unique_products_df)} rows")
    
    if asp_df.empty or volume_df.empty or unique_products_df.empty:
        print("No data returned. Exiting.")
        return
    
    # Pivot ASP data - weeks as columns
    asp_pivot = asp_df.pivot(index='condition', columns='week_start', values='avg_market_price')
    asp_pivot = asp_pivot.round(2)
    asp_pivot.columns = [f"{col.strftime('%Y-%m-%d')}" for col in asp_pivot.columns]
    asp_pivot = asp_pivot.reset_index()
    asp_pivot.columns.name = None
    
    # Pivot Volume data - weeks as columns
    volume_pivot = volume_df.pivot(index='condition', columns='week_start', values='quantity_sold')
    volume_pivot = volume_pivot.fillna(0).astype(int)
    volume_pivot.columns = [f"{col.strftime('%Y-%m-%d')}" for col in volume_pivot.columns]
    volume_pivot = volume_pivot.reset_index()
    volume_pivot.columns.name = None
    
    # Pivot Unique Products data - weeks as columns
    unique_products_pivot = unique_products_df.pivot(index='metric', columns='week_start', values='unique_products_with_sales')
    unique_products_pivot = unique_products_pivot.fillna(0).astype(int)
    unique_products_pivot.columns = [f"{col.strftime('%Y-%m-%d')}" for col in unique_products_pivot.columns]
    unique_products_pivot = unique_products_pivot.reset_index()
    unique_products_pivot.columns.name = None
    
    # Create output directory if it doesn't exist
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create formatted Excel file
    excel_file = f"{output_dir}/{timestamp}_bda_weekly_summary_formatted.xlsx"
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Summary"
    
    # Add formatted data sections
    current_row = 1
    
    # ASP Section
    current_row = add_formatted_data(ws, asp_pivot, current_row, "Weekly Average Selling Price (ASP) by Condition")
    
    # Volume Section  
    current_row = add_formatted_data(ws, volume_pivot, current_row, "Weekly Quantity Sold by Condition")
    
    # Unique Products Section
    current_row = add_formatted_data(ws, unique_products_pivot, current_row, "Weekly Number of Unique Products with Sales")
    
    wb.save(excel_file)
    
    print(f"Formatted Excel file saved: {excel_file}")
    print(f"ASP Data Shape: {asp_pivot.shape}")
    print(f"Volume Data Shape: {volume_pivot.shape}")
    print(f"Unique Products Data Shape: {unique_products_pivot.shape}")
    
    print("\nConditions in data:")
    print("ASP:", asp_pivot['condition'].tolist())
    print("Volume:", volume_pivot['condition'].tolist())
    print("Unique Products:", unique_products_pivot['metric'].tolist())

if __name__ == "__main__":
    main()