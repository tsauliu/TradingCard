#!/usr/bin/env python3
"""
Create PSA auction price weekly analysis
Weekly average prices by PSA level and card name with smart outlier detection
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def detect_outliers(df, price_col='auction_price', window_size=5, threshold=0.75):
    """
    Multi-pass outlier detection with IQR and extreme outlier handling
    
    Args:
        df: DataFrame with price data
        price_col: Column name for prices
        window_size: Size of rolling window for median calculation
        threshold: Threshold for rolling median detection (0.75 = 75%)
    
    Returns:
        DataFrame with outliers removed
    """
    df = df.copy()
    df = df.sort_values(['card_name', 'psa_level', 'auction_date']).reset_index(drop=True)
    
    total_count = len(df)
    outlier_counts = {'extreme': 0, 'iqr': 0, 'rolling': 0}
    
    # Group by card_name and psa_level for outlier detection
    def remove_outliers_group(group):
        nonlocal outlier_counts
        if len(group) < 3:  # Need at least 3 points for meaningful outlier detection
            return group
        
        group = group.copy().reset_index(drop=True)
        prices = group[price_col].values
        
        # Pass 1: Remove extreme outliers (>5x median)
        group_median = np.median(prices)
        extreme_mask = prices > (5 * group_median)
        outlier_counts['extreme'] += extreme_mask.sum()
        
        if extreme_mask.any():
            group = group[~extreme_mask].reset_index(drop=True)
            prices = group[price_col].values
        
        if len(group) < 3:
            return group
        
        # Pass 2: IQR-based outlier detection
        Q1 = np.percentile(prices, 25)
        Q3 = np.percentile(prices, 75)
        IQR = Q3 - Q1
        
        if IQR > 0:  # Only apply if there's variance
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            iqr_mask = (prices < lower_bound) | (prices > upper_bound)
            outlier_counts['iqr'] += iqr_mask.sum()
            
            if iqr_mask.any():
                group = group[~iqr_mask].reset_index(drop=True)
                prices = group[price_col].values
        
        if len(group) < window_size:
            return group
        
        # Pass 3: Rolling median with tighter threshold
        rolling_median = pd.Series(prices).rolling(window=window_size, center=True, min_periods=3).median()
        rolling_median = rolling_median.fillna(np.median(prices))
        
        # Identify outliers with tighter threshold
        outliers_mask = np.abs(prices - rolling_median) > (threshold * rolling_median)
        
        # Improved trend detection: require 5+ consecutive points and reasonable variance
        trend_mask = np.zeros_like(outliers_mask, dtype=bool)
        min_trend_length = 5
        
        for i in range(len(outliers_mask) - min_trend_length + 1):
            trend_slice = outliers_mask[i:i+min_trend_length]
            if trend_slice.all():
                # Check if trend values are reasonably close to each other
                trend_prices = prices[i:i+min_trend_length]
                trend_cv = np.std(trend_prices) / np.mean(trend_prices)  # Coefficient of variation
                
                # Only preserve trend if coefficient of variation is < 0.5 (reasonable consistency)
                if trend_cv < 0.5:
                    trend_mask[i:i+min_trend_length] = True
        
        # Remove outliers that aren't part of legitimate trends
        final_outliers = outliers_mask & ~trend_mask
        outlier_counts['rolling'] += final_outliers.sum()
        
        return group[~final_outliers]
    
    cleaned_df = df.groupby(['card_name', 'psa_level']).apply(remove_outliers_group).reset_index(drop=True)
    
    total_removed = sum(outlier_counts.values())
    print(f"Multi-pass outlier detection complete:")
    print(f"  Original records: {total_count:,}")
    print(f"  Extreme outliers removed (>5x median): {outlier_counts['extreme']:,}")
    print(f"  IQR outliers removed: {outlier_counts['iqr']:,}")
    print(f"  Rolling median outliers removed: {outlier_counts['rolling']:,}")
    print(f"  Total outliers removed: {total_removed:,}")
    print(f"  Records remaining: {len(cleaned_df):,}")
    print(f"  Removal rate: {total_removed/total_count*100:.1f}%")
    
    return cleaned_df

def add_nested_formatted_data(ws, df, start_row, title):
    """Add nested formatted data to worksheet with PSA level in A4 format, card names in B5 format"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add data starting from next row
    data_start_row = start_row + 2
    current_row = data_start_row
    
    # Add headers for card-first nested structure
    headers = ['Card Name', 'PSA Level'] + [col for col in df.columns if col not in ['psa_level', 'card_name']]
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    current_row += 1
    last_psa_level = None
    
    # Data is already sorted (PSA level descending, then card name)
    df_sorted = df.copy()
    
    # Group by card name first, then nest PSA levels
    for card_name in df_sorted['card_name'].unique():
        # Add card name header row spanning all columns
        card_header_cell = ws.cell(row=current_row, column=1, value=card_name)
        card_header_cell.font = Font(bold=True, size=11, color="FFFFFF")
        card_header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        card_header_cell.border = thin_border
        
        # Fill remaining cells in card header row
        for col_idx in range(2, len(headers) + 1):
            header_cell = ws.cell(row=current_row, column=col_idx, value="")
            header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_cell.border = thin_border
        
        current_row += 1
        
        # Add PSA level rows for this card (sorted by PSA level descending)
        card_psa_data = df_sorted[df_sorted['card_name'] == card_name].sort_values('psa_level', ascending=False, key=lambda x: pd.to_numeric(x, errors='coerce'))
        for _, row in card_psa_data.iterrows():
            psa_level = row['psa_level']
            
            # Column A: Empty (since card name is in header row above)
            empty_cell = ws.cell(row=current_row, column=1, value="")
            empty_cell.border = thin_border
            
            # Column B: PSA Level
            psa_cell = ws.cell(row=current_row, column=2, value=f"PSA {psa_level}")
            psa_cell.border = thin_border
            
            # Add data for each column starting from column C
            col_idx = 3
            for col in headers[2:]:
                if col in row.index:
                    value = row[col]
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # Format different column types
                    if col == 'Lifecycle Sales' and isinstance(value, (int, float)) and pd.notnull(value):
                        cell.number_format = '#,##0'
                    elif col == 'PSA URL' and value:
                        # Keep URL as text, no special formatting needed
                        pass
                    elif col not in ['Lifecycle Sales', 'PSA URL'] and isinstance(value, (int, float)) and pd.notnull(value):
                        # Price columns - format as currency
                        cell.number_format = '$#,##0.00'
                else:
                    cell = ws.cell(row=current_row, column=col_idx, value="")
                    cell.border = thin_border
                
                col_idx += 1
            
            # Alternate row colors for PSA level rows only
            if current_row % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if cell.fill.start_color.rgb not in ["366092", "FF366092"]:
                        cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            
            current_row += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in this column
        for check_row in range(data_start_row, current_row):
            cell_value = ws.cell(row=check_row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 35)
    
    return current_row + 2

def add_simple_formatted_data(ws, df, start_row, title):
    """Add simple formatted data to worksheet for single PSA level"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add data starting from next row
    data_start_row = start_row + 2
    current_row = data_start_row
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    current_row += 1
    
    # Add data rows (sorted by card order already from main pivot)
    for _, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format different column types
            col_name = df.columns[col_idx-1]
            if col_name == 'Price Points' and isinstance(value, (int, float)) and pd.notnull(value):
                cell.number_format = '#,##0'
            elif col_name == 'Lifecycle Sales' and isinstance(value, (int, float)) and pd.notnull(value):
                cell.number_format = '#,##0'
            elif col_name == 'PSA URL' and value:
                # Keep URL as text, no special formatting needed
                pass
            elif col_name not in ['card_name', 'Price Points', 'Lifecycle Sales', 'PSA URL'] and isinstance(value, (int, float)) and pd.notnull(value):
                # Price columns - format as currency
                cell.number_format = '$#,##0.00'
            
            # Alternate row colors
            if current_row % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        current_row += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in this column
        for check_row in range(data_start_row, current_row):
            cell_value = ws.cell(row=check_row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 35)
    
    return current_row + 2

def main():
    # Set up credentials
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/Analysis/service-account.json'
    client = bigquery.Client()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    # Query to pull PSA auction price data with lifecycle sales and URL
    query = """
    WITH sales_data AS (
      SELECT 
          card_name,
          grade as psa_level,
          sale_price as auction_price,
          psa_url,
          PARSE_DATE('%m/%d/%Y', sale_date) as auction_date,
          DATE_TRUNC(PARSE_DATE('%m/%d/%Y', sale_date), WEEK(MONDAY)) as week_start
      FROM `rising-environs-456314-a3.tcg_data.psa_auction_prices`
      WHERE record_type = 'sale'
        AND sale_price IS NOT NULL 
        AND sale_price > 0
        AND card_name IS NOT NULL
        AND card_name != ''
        AND grade IS NOT NULL
        AND sale_date IS NOT NULL
        AND sale_date != ''
        AND PARSE_DATE('%m/%d/%Y', sale_date) >= '2016-01-01'
    ),
    summary_data AS (
      SELECT 
          card_name,
          grade as psa_level,
          total_sales_count,
          psa_url
      FROM `rising-environs-456314-a3.tcg_data.psa_auction_prices`
      WHERE record_type = 'summary'
        AND total_sales_count IS NOT NULL
    )
    SELECT 
        s.card_name,
        s.psa_level,
        s.auction_price,
        COALESCE(sum.total_sales_count, 0) as total_sales_count,
        COALESCE(s.psa_url, sum.psa_url) as psa_url,
        s.auction_date,
        s.week_start
    FROM sales_data s
    LEFT JOIN summary_data sum ON s.card_name = sum.card_name AND s.psa_level = sum.psa_level
    ORDER BY s.card_name, s.psa_level, s.auction_date
    """
    
    print("Executing PSA auction price query...")
    try:
        df = client.query(query).to_dataframe()
        print(f"Query completed. Retrieved {len(df):,} rows.")
    except Exception as e:
        print(f"Error executing query: {e}")
        return
    
    if df.empty:
        print("No data returned. Exiting.")
        return
    
    print(f"\nData summary:")
    print(f"  Unique cards: {df['card_name'].nunique():,}")
    print(f"  PSA levels: {sorted(df['psa_level'].unique())}")
    print(f"  Price range: ${df['auction_price'].min():.2f} - ${df['auction_price'].max():,.2f}")
    print(f"  Date range: {df['auction_date'].min()} to {df['auction_date'].max()}")
    
    # Skip outlier detection - use raw data
    print(f"\nUsing raw data without outlier detection...")
    cleaned_df = df
    
    # Create weekly aggregation
    print(f"Creating weekly aggregation...")
    weekly_data = cleaned_df.groupby(['card_name', 'psa_level', 'week_start']).agg({
        'auction_price': ['mean', 'count'],
        'total_sales_count': 'first',  # Take first value since it should be same for each card/grade combo
        'psa_url': 'first'  # Take first URL for each card/grade combo
    }).reset_index()
    
    # Flatten column names
    weekly_data.columns = ['card_name', 'psa_level', 'week_start', 'avg_price', 'transaction_count', 'total_sales_count', 'psa_url']
    weekly_data['avg_price'] = weekly_data['avg_price'].round(2)
    
    print(f"Weekly aggregation complete:")
    print(f"  Unique card/PSA combinations: {weekly_data.groupby(['card_name', 'psa_level']).ngroups:,}")
    print(f"  Weeks covered: {weekly_data['week_start'].nunique()}")
    
    # Use raw weekly data without creating artificial weeks
    print(f"Using only weeks with actual sales data...")
    complete_weekly_data = weekly_data
    
    # Create pivot table with weeks as columns
    print(f"Creating nested pivot table...")
    
    # Calculate number of weeks with price data for each card/PSA combination
    weeks_with_prices = complete_weekly_data[complete_weekly_data['avg_price'].notnull()].groupby(['card_name', 'psa_level']).size().reset_index(name='weeks_with_prices')
    
    # Calculate price point count by card/PSA combination (for PSA level ranking within cards)
    card_psa_price_points = df.groupby(['card_name', 'psa_level']).size().reset_index(name='psa_price_point_count')
    
    # Calculate price point count by card (for summary display)
    card_price_points = df.groupby('card_name').size().reset_index(name='card_price_point_count')
    
    # Calculate total sales count by card for ranking cards
    card_total_sales = complete_weekly_data.groupby('card_name')['total_sales_count'].first().reset_index()
    # Only drop rows where total_sales_count is actually null (not 0)
    card_total_sales = card_total_sales[card_total_sales['total_sales_count'].notnull()]
    
    # Sort cards by lifecycle sales (as before)
    card_total_sales = card_total_sales.sort_values('total_sales_count', ascending=False)
    card_order = {card: idx for idx, card in enumerate(card_total_sales['card_name'])}
    
    # Create card to total sales mapping for adding to output
    card_sales_map = dict(zip(card_total_sales['card_name'], card_total_sales['total_sales_count']))
    
    # Create card/PSA to weeks with prices mapping for adding to output
    weeks_with_prices_map = {}
    for _, row in weeks_with_prices.iterrows():
        weeks_with_prices_map[(row['card_name'], row['psa_level'])] = row['weeks_with_prices']
    
    # Create card/PSA to price point count mapping for PSA level ranking
    card_psa_price_points_map = {}
    for _, row in card_psa_price_points.iterrows():
        card_psa_price_points_map[(row['card_name'], row['psa_level'])] = row['psa_price_point_count']
    
    # Convert PSA level to numeric for proper sorting, handle string grades
    def convert_psa_level(psa_str):
        try:
            return float(psa_str)
        except:
            return 0.0  # Handle any non-numeric grades
    
    complete_weekly_data['psa_numeric'] = complete_weekly_data['psa_level'].apply(convert_psa_level)
    complete_weekly_data['card_order'] = complete_weekly_data['card_name'].map(card_order)
    
    # Create pivot table with multi-index (psa_level, card_name)
    pivot_df = complete_weekly_data.pivot_table(
        index=['card_order', 'psa_numeric', 'psa_level', 'card_name'], 
        columns='week_start', 
        values='avg_price',
        aggfunc='mean'
    )
    
    # Sort by card order (lifecycle sales), then by PSA level descending (10 to 0)
    pivot_df = pivot_df.sort_index(level=['card_order', 'psa_numeric'], ascending=[True, False])
    
    # Drop the helper columns from index
    pivot_df = pivot_df.droplevel(['card_order', 'psa_numeric'])
    
    pivot_df = pivot_df.round(2)
    
    # Format column names as dates
    pivot_df.columns = [col.strftime('%Y-%m-%d') if pd.notnull(col) else 'Unknown' for col in pivot_df.columns]
    
    # Reset index to get psa_level and card_name as columns
    pivot_df = pivot_df.reset_index()
    pivot_df.columns.name = None
    
    # Add weeks with prices count, lifecycle sales count and PSA URL as the first data columns
    pivot_df['Price Points'] = pivot_df.apply(lambda row: weeks_with_prices_map.get((row['card_name'], row['psa_level']), 0), axis=1)
    pivot_df['Lifecycle Sales'] = pivot_df['card_name'].map(card_sales_map)
    
    # Add PSA URL (get first non-null URL for each card/PSA combo)
    card_psa_url_map = complete_weekly_data.groupby(['card_name', 'psa_level'])['psa_url'].first().to_dict()
    pivot_df['PSA URL'] = pivot_df.apply(lambda row: card_psa_url_map.get((row['card_name'], row['psa_level']), ''), axis=1)
    
    # Reorder columns to have Price Points, Lifecycle Sales and URL first after card info
    date_cols = [col for col in pivot_df.columns if col not in ['psa_level', 'card_name', 'Price Points', 'Lifecycle Sales', 'PSA URL']]
    pivot_df = pivot_df[['psa_level', 'card_name', 'Price Points', 'Lifecycle Sales', 'PSA URL'] + date_cols]
    
    print(f"Nested pivot table created: {pivot_df.shape[0]} rows x {pivot_df.shape[1]} columns")
    print(f"  Raw data only - no forward-fill applied")
    
    # Create output directory if it doesn't exist
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create formatted Excel file
    excel_file = f"{output_dir}/{timestamp}_psa_weekly_analysis.xlsx"
    
    # Create workbook with multiple sheets
    wb = Workbook()
    
    # Main sheet with all PSA levels
    ws1 = wb.active
    ws1.title = "All PSA Levels"
    current_row = 1
    current_row = add_nested_formatted_data(ws1, pivot_df, current_row, "Weekly Average PSA Auction Prices by Grade and Card")
    
    # Create PSA 9 only sheet
    ws2 = wb.create_sheet(title="PSA 9 Only")
    
    # Filter for PSA 9 data only and simplify for single PSA level
    psa9_pivot = pivot_df[pivot_df['psa_level'] == '9'].copy()
    
    if not psa9_pivot.empty:
        # Remove PSA level column since it's all PSA 9
        psa9_simple = psa9_pivot.drop('psa_level', axis=1).copy()
        # Reorder to put card_name first
        date_cols = [col for col in psa9_simple.columns if col not in ['card_name', 'Price Points', 'Lifecycle Sales', 'PSA URL']]
        psa9_simple = psa9_simple[['card_name', 'Price Points', 'Lifecycle Sales', 'PSA URL'] + date_cols]
        
        # Add simple formatted data (no nesting needed for single PSA level)
        current_row = 1
        current_row = add_simple_formatted_data(ws2, psa9_simple, current_row, "PSA 9 Weekly Average Auction Prices")
    else:
        # Add message if no PSA 9 data
        ws2.cell(row=1, column=1, value="No PSA 9 data available")
    
    # Save workbook
    wb.save(excel_file)
    
    print(f"\nExcel file saved: {excel_file}")
    print(f"  Sheet 1 'All PSA Levels': {pivot_df.shape}")
    if not psa9_pivot.empty:
        print(f"  Sheet 2 'PSA 9 Only': {psa9_pivot.shape[0]} cards x {psa9_simple.shape[1]} columns")
    else:
        print(f"  Sheet 2 'PSA 9 Only': No PSA 9 data available")
    
    # Summary statistics
    print(f"\nSummary statistics:")
    print(f"  Total card/PSA combinations: {len(pivot_df)}")
    print(f"  Weeks of data: {len([col for col in pivot_df.columns if col not in ['psa_level', 'card_name']])}")
    print(f"  Average transactions per week: {weekly_data['transaction_count'].mean():.1f}")
    
    # Show cards ranked by lifecycle sales (as they appear in output)
    print(f"\nCards ranked by lifecycle sales (as shown in output):")
    for _, row in card_total_sales.head(10).iterrows():
        # Get total price points for this card
        card_price_point_total = card_price_points[card_price_points['card_name'] == row['card_name']]['card_price_point_count'].iloc[0] if not card_price_points[card_price_points['card_name'] == row['card_name']].empty else 0
        print(f"  {row['card_name']}: {row['total_sales_count']:,} lifecycle sales, {card_price_point_total:,} price points")
    
    print(f"\nNote: Within each card, PSA levels are ordered by grade (10, 9, 8, 7... 0)")
    print(f"'Price Points' column shows number of weeks with actual price data for each PSA level")
    
    # Show top 10 most valuable combinations
    print(f"\nTop 10 highest average prices (most recent week):")
    price_cols = [col for col in pivot_df.columns if col not in ['psa_level', 'card_name']]
    if price_cols:
        latest_week_col = price_cols[-1]
        pivot_df['combo'] = 'PSA ' + pivot_df['psa_level'].astype(str) + ' - ' + pivot_df['card_name']
        top_cards = pivot_df[['combo', latest_week_col]].dropna().nlargest(10, latest_week_col)
        for _, row in top_cards.iterrows():
            print(f"  {row['combo']}: ${row[latest_week_col]:,.2f}")

if __name__ == "__main__":
    main()