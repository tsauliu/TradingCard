#!/usr/bin/env python3
"""
Top 50 Products Analysis from tcg_prices_bda
Creates 2-sheet Excel: Summary and Detailed data
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def add_formatted_data(ws, df, start_row, title, price_format=False):
    """Add formatted data to worksheet"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add data starting from next row
    data_start_row = start_row + 2
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Add data rows
    last_data_row = data_start_row
    for row_idx, row in enumerate(df.itertuples(index=False), data_start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, (int, float)) and col_idx > 1:
                if price_format:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        last_data_row = row_idx
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check header length
        header_value = ws.cell(row=data_start_row, column=col_idx).value
        if header_value:
            max_length = max(max_length, len(str(header_value)))
        
        # Check data lengths
        for check_row in range(data_start_row + 1, last_data_row + 1):
            cell_value = ws.cell(row=check_row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
    
    return last_data_row + 3

def main():
    # Set up credentials
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/5_Analysis/service-account.json'
    client = bigquery.Client()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    # Step 1: Get top 200 products by lifecycle quantity sold (ASP > $5, first week qty > 0) with all metadata
    top_products_query = """
    WITH first_week_products AS (
      SELECT 
        product_id,
        MIN(DATE_TRUNC(bucket_start_date, WEEK(MONDAY))) as first_week
      FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda`
      WHERE bucket_start_date IS NOT NULL 
        AND quantity_sold IS NOT NULL
      GROUP BY product_id
    ),
    first_week_quantities AS (
      SELECT 
        p.product_id,
        SUM(p.quantity_sold) as first_week_quantity
      FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` p
      INNER JOIN first_week_products fw 
        ON p.product_id = fw.product_id 
        AND DATE_TRUNC(p.bucket_start_date, WEEK(MONDAY)) = fw.first_week
      WHERE p.quantity_sold IS NOT NULL
      GROUP BY p.product_id
      HAVING SUM(p.quantity_sold) > 0
    )
    SELECT 
        p.product_id,
        SUM(p.quantity_sold) as lifecycle_quantity_sold,
        CASE 
          WHEN SUM(p.quantity_sold) > 0 
          THEN SUM(p.quantity_sold * p.market_price) / SUM(p.quantity_sold)
          ELSE 0 
        END as lifecycle_asp,
        m.category_categoryId,
        m.category_name,
        m.category_displayName,
        m.category_seoCategoryName,
        m.category_categoryDescription,
        m.category_categoryPageTitle,
        m.category_sealedLabel,
        m.category_nonSealedLabel,
        m.category_conditionGuideUrl,
        m.category_isScannable,
        m.category_popularity,
        m.category_isDirect,
        m.group_groupId,
        m.group_name,
        m.group_abbreviation,
        m.group_isSupplemental,
        m.group_publishedOn,
        m.group_categoryId,
        m.product_productId,
        m.product_name,
        m.product_cleanName,
        m.product_imageUrl,
        m.product_categoryId,
        m.product_groupId,
        m.product_url,
        m.product_modifiedOn,
        m.product_imageCount,
        m.product_presaleInfo,
        m.product_extendedData
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` p
    INNER JOIN first_week_quantities fwq 
      ON p.product_id = fwq.product_id
    LEFT JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` m
      ON CAST(p.product_id AS STRING) = CAST(m.product_productId AS STRING)
    WHERE p.quantity_sold IS NOT NULL 
      AND p.market_price > 0
    GROUP BY p.product_id, m.category_categoryId, m.category_name, m.category_displayName, m.category_seoCategoryName, m.category_categoryDescription, m.category_categoryPageTitle, m.category_sealedLabel, m.category_nonSealedLabel, m.category_conditionGuideUrl, m.category_isScannable, m.category_popularity, m.category_isDirect, m.group_groupId, m.group_name, m.group_abbreviation, m.group_isSupplemental, m.group_publishedOn, m.group_categoryId, m.product_productId, m.product_name, m.product_cleanName, m.product_imageUrl, m.product_categoryId, m.product_groupId, m.product_url, m.product_modifiedOn, m.product_imageCount, m.product_presaleInfo, m.product_extendedData
    HAVING CASE 
             WHEN SUM(p.quantity_sold) > 0 
             THEN SUM(p.quantity_sold * p.market_price) / SUM(p.quantity_sold)
             ELSE 0 
           END > 5
    ORDER BY lifecycle_quantity_sold DESC
    LIMIT 200
    """
    
    print("Getting top 200 products by lifecycle quantity sold (ASP > $5, first week qty > 0)...")
    top_products_df = client.query(top_products_query).to_dataframe()
    
    if top_products_df.empty:
        print("No products found. Exiting.")
        return
    
    print(f"Found {len(top_products_df)} top products")
    
    # Get the list of top 200 product IDs
    top_product_ids = top_products_df['product_id'].tolist()
    top_product_ids_str = "', '".join([str(pid) for pid in top_product_ids])
    
    # Step 2: Get weekly data for these top 200 products
    weekly_data_query = f"""
    SELECT 
        p.product_id,
        m.product_name,
        p.condition,
        DATE_TRUNC(p.bucket_start_date, WEEK(MONDAY)) as week_start,
        AVG(p.market_price) as avg_market_price,
        SUM(p.quantity_sold) as weekly_quantity_sold
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` p
    LEFT JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` m
      ON CAST(p.product_id AS STRING) = CAST(m.product_productId AS STRING)
    WHERE p.product_id IN ('{top_product_ids_str}')
      AND p.bucket_start_date IS NOT NULL
      AND p.bucket_start_date >= '2024-01-01'
      AND p.condition IS NOT NULL
      AND p.condition != ''
      AND p.market_price > 0
      AND p.quantity_sold IS NOT NULL
    GROUP BY p.product_id, m.product_name, p.condition, DATE_TRUNC(p.bucket_start_date, WEEK(MONDAY))
    ORDER BY p.product_id, week_start, p.condition
    """
    
    print("Getting weekly data for top 200 products...")
    weekly_data_df = client.query(weekly_data_query).to_dataframe()
    
    if weekly_data_df.empty:
        print("No weekly data found. Exiting.")
        return
    
    print(f"Retrieved {len(weekly_data_df)} weekly data records")
    
    # Step 3: Calculate weighted averages and totals by week and product
    print("Calculating weighted averages...")
    
    # Calculate weighted price and total quantity per product per week
    weekly_summary = []
    
    for product_id in top_product_ids:
        product_data = weekly_data_df[weekly_data_df['product_id'] == product_id]
        
        if product_data.empty:
            continue
        
        # Get product name (should be same for all rows of this product)
        product_name = product_data['product_name'].iloc[0] if not product_data['product_name'].isna().all() else f"Product_{product_id}"
            
        for week in product_data['week_start'].unique():
            week_data = product_data[product_data['week_start'] == week]
            
            # Calculate weighted average price
            total_value = (week_data['avg_market_price'] * week_data['weekly_quantity_sold']).sum()
            total_quantity = week_data['weekly_quantity_sold'].sum()
            
            if total_quantity > 0:
                weighted_price = total_value / total_quantity
            else:
                weighted_price = 0
            
            weekly_summary.append({
                'product_id': product_id,
                'product_name': product_name,
                'week_start': week,
                'weighted_avg_price': round(weighted_price, 2),
                'total_weekly_quantity': int(total_quantity)
            })
    
    weekly_summary_df = pd.DataFrame(weekly_summary)
    
    # Step 4: Create 3 separate datasets - Metadata, Prices, Quantities
    print("Creating 3 separate datasets: Metadata, Prices, Quantities...")
    
    # Ensure consistent product_id ordering across all sheets
    product_order = top_products_df['product_id'].tolist()
    
    # 1. Metadata sheet - only specified columns
    metadata_columns = ['product_id', 'lifecycle_quantity_sold', 'group_name', 'product_cleanName', 'product_url']
    metadata_df = top_products_df[metadata_columns].copy()
    
    # Create new combined column
    metadata_df['group_product_id'] = (
        metadata_df['group_name'].fillna('').astype(str) + ' ' + 
        metadata_df['product_cleanName'].fillna('').astype(str) + ' ' + 
        metadata_df['product_id'].astype(str)
    )
    
    # 2. Create price dataset
    price_data_list = []
    for product_id in product_order:
        row_data = {'product_id': product_id}
        product_weekly = weekly_summary_df[weekly_summary_df['product_id'] == product_id]
        
        for week in sorted(weekly_summary_df['week_start'].unique()):
            week_data = product_weekly[product_weekly['week_start'] == week]
            if not week_data.empty:
                row_data[week] = round(week_data['weighted_avg_price'].iloc[0], 2)
            else:
                row_data[week] = 0.0
        price_data_list.append(row_data)
    
    prices_df = pd.DataFrame(price_data_list)
    
    # 3. Create quantity dataset
    quantity_data_list = []
    for product_id in product_order:
        row_data = {'product_id': product_id}
        product_weekly = weekly_summary_df[weekly_summary_df['product_id'] == product_id]
        
        for week in sorted(weekly_summary_df['week_start'].unique()):
            week_data = product_weekly[product_weekly['week_start'] == week]
            if not week_data.empty:
                row_data[week] = int(week_data['total_weekly_quantity'].iloc[0])
            else:
                row_data[week] = 0
        quantity_data_list.append(row_data)
    
    quantities_df = pd.DataFrame(quantity_data_list)
    
    # Create output directory if it doesn't exist
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create Excel file with 3 separate sheets
    excel_file = f"{output_dir}/{timestamp}_top200_products_analysis.xlsx"
    
    wb = Workbook()
    
    # Sheet 1: Metadata
    ws_metadata = wb.active
    ws_metadata.title = "Metadata"
    add_formatted_data(ws_metadata, metadata_df, 1, 
                      "Top 200 Products - Metadata, Quantity Sold & ASP", price_format=False)
    
    # Sheet 2: Prices
    ws_prices = wb.create_sheet(title="Prices")
    add_formatted_data(ws_prices, prices_df, 1, 
                      "Top 200 Products - Weekly Prices", price_format=True)
    
    # Sheet 3: Quantities
    ws_quantities = wb.create_sheet(title="Quantities")
    add_formatted_data(ws_quantities, quantities_df, 1, 
                      "Top 200 Products - Weekly Quantities", price_format=False)
    
    # Optional: Keep detailed sheet as well
    ws_detailed = wb.create_sheet(title="Detailed Weekly by Condition")
    
    # Format detailed data
    detailed_df = weekly_data_df.copy()
    if pd.api.types.is_datetime64_any_dtype(detailed_df['week_start']):
        detailed_df['week_start'] = detailed_df['week_start'].dt.strftime('%Y-%m-%d')
    detailed_df = detailed_df.round({'avg_market_price': 2})
    
    add_formatted_data(ws_detailed, detailed_df, 1, "Detailed Weekly Data - All Products by Condition")
    
    wb.save(excel_file)
    
    print(f"Excel file saved: {excel_file}")
    print(f"Metadata sheet shape: {metadata_df.shape}")
    print(f"Prices sheet shape: {prices_df.shape}")
    print(f"Quantities sheet shape: {quantities_df.shape}")
    print(f"Detailed data shape: {detailed_df.shape}")
    
    print(f"\nTop 10 products by lifecycle quantity sold (ASP > $5, first week qty > 0):")
    print(top_products_df.head(10)[['product_id', 'product_name', 'lifecycle_quantity_sold', 'lifecycle_asp']])

if __name__ == "__main__":
    main()