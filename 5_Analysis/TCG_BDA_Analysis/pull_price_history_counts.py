#!/usr/bin/env python3
"""
Pull price history counts by category_id and group_id from tcg_prices
Join with tcg_metadata to get category and group names
Export to formatted Excel
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_formatted_excel(df, filename):
    """Create formatted Excel file with professional styling"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Price History Counts"
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Add data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, (int, float)) and col_idx == 2:  # total_price_records column
                cell.number_format = '#,##0'
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Check header length
        header_value = ws.cell(row=1, column=col_idx).value
        if header_value:
            max_length = max(max_length, len(str(header_value)))
        
        # Check data lengths (sample first 100 rows for performance)
        for check_row in range(2, min(102, len(df) + 2)):
            cell_value = ws.cell(row=check_row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(filename)

def main():
    # Set up credentials
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/5_Analysis/service-account.json'
    client = bigquery.Client()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    # Load Categories.csv for category name lookup
    categories_df = pd.read_csv('/home/caoliu/TradingCard/5_Analysis/TCG_BDA_Analysis/Categories.csv')
    categories_lookup = dict(zip(categories_df['categoryId'], categories_df['name']))
    
    # Create eBay-friendly search names mapping - aggressively combine similar terms
    ebay_search_mapping = {
        'Magic': 'magic the gathering card',
        'YuGiOh': 'yu-gi-oh card',
        'Pokemon': 'pokemon card',
        'Pokemon Japan': 'pokemon card',
        'Weiss Schwarz': 'weiss schwarz card',
        'Cardfight Vanguard': 'cardfight vanguard card',
        'Dragon Ball Super CCG': 'dragon ball card',
        'Dragon Ball Super Fusion World': 'dragon ball card',
        'Dragon Ball Z TCG': 'dragon ball card',
        'Force of Will': 'force of will card',
        'Flesh & Blood TCG': 'flesh and blood card',
        'Flesh and Blood TCG': 'flesh and blood card',
        'Future Card BuddyFight': 'buddyfight card',
        'Final Fantasy TCG': 'final fantasy card',
        'UniVersus': 'universus card',
        'Digimon Card Game': 'digimon card',
        'WoW': 'world of warcraft card',
        'Heroclix': 'heroclix figure',
        'Card Sleeves': 'card sleeves',
        'MetaZoo': 'metazoo card',
        'WIXOSS': 'wixoss card',
        'One Piece Card Game': 'one piece card',
        'Grand Archive': 'grand archive card',
        'Star Wars Unlimited': 'star wars card',
        'Lorcana TCG': 'disney lorcana card',
        'Disney Lorcana': 'disney lorcana card',
        'Shadowverse Evolve': 'shadowverse card',
        'Battle Spirits Saga': 'battle spirits card',
        'Playmats': 'playmat',
        'Sorcery Contested Realm': 'sorcery card',
        'Sorcery: Contested Realm': 'sorcery card',
        'Deck Boxes': 'deck box',
        'Star Wars Destiny': 'star wars card',
        'Star Wars: Destiny': 'star wars card',
        'Dice Masters': 'dice masters card',
        'Akora': 'akora card',
        'Alpha Clash': 'alpha clash card',
        'Gate Ruler': 'gate ruler card',
        'Kryptik TCG': 'kryptik card',
        'Argent Saga TCG': 'argent saga card',
        'Bakugan TCG': 'bakugan card',
        'Lightseekers TCG': 'lightseekers card',
        'Warhammer Age of Sigmar Champions TCG': 'warhammer card',
        'Union Arena': 'union arena card',
        'MetaX TCG': 'metax card',
        'D & D Miniatures': 'dungeons and dragons miniature',
        'The Caster Chronicles': 'caster chronicles card',
        'Elestrals': 'elestrals card',
        'Transformers TCG': 'transformers card',
        'Life Counters': 'life counter',
        'Funko': 'funko pop',
        'Munchkin CCG': 'munchkin card',
        'Storage Albums': 'card binder',
        'Collectible Storage': 'card storage',
        'Dragoborne': 'dragoborne card',
        'Exodus TCG': 'exodus card',
        'Chrono Clash System': 'chrono clash card',
        'Bulk Lots': 'card bulk lot',
        'Gundam Card Game': 'gundam card',
        'Protective Pages': 'card pages',
        'Zombie World Order TCG': 'zombie world order card',
        'KeyForge': 'keyforge card',
        'My Little Pony CCG': 'my little pony card',
        'hololive OFFICIAL CARD GAME': 'hololive card',
        'Godzilla Card Game': 'godzilla card',
        'Alternate Souls': 'alternate souls card',
        'Riftbound League of Legends Trading Card Game': 'league of legends card',
        'Riftbound: League of Legends Trading Card Game': 'league of legends card'
    }
    
    # SQL query to count price records by category_id only
    query = """
    SELECT 
      category_id,
      COUNT(*) as total_price_records
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices`
    GROUP BY category_id
    ORDER BY COUNT(*) DESC
    """
    
    print("Executing query to pull price history counts by category...")
    
    # Execute query
    query_job = client.query(query)
    df = query_job.result().to_dataframe()
    
    # Add category names from Categories.csv
    df['category_name'] = df['category_id'].map(categories_lookup)
    
    # Add eBay-friendly search names
    df['ebay_search_name'] = df['category_name'].map(ebay_search_mapping)
    
    # For categories not in the mapping, use the original category_name as fallback
    df['ebay_search_name'] = df['ebay_search_name'].fillna(df['category_name'])
    
    print(f"Query completed. Retrieved {len(df)} categories.")
    
    if df.empty:
        print("No data returned. Exiting.")
        return
    
    # Create output directory if it doesn't exist
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create formatted Excel file
    excel_file = f"{output_dir}/{timestamp}_category_price_history_counts.xlsx"
    create_formatted_excel(df, excel_file)
    
    print(f"Formatted Excel file saved: {excel_file}")
    
    # Display summary statistics
    print(f"\nSummary:")
    print(f"Total categories: {len(df)}")
    print(f"Total price records across all categories: {df['total_price_records'].sum():,}")
    print(f"Average price records per category: {df['total_price_records'].mean():.0f}")
    print(f"Max price records for single category: {df['total_price_records'].max():,}")
    print(f"Min price records for single category: {df['total_price_records'].min():,}")
    
    # Show all categories by record count
    print(f"\nAll categories by price record count:")
    print(df[['category_name', 'ebay_search_name', 'total_price_records']].to_string(index=False))
    
    return df

if __name__ == "__main__":
    data = main()