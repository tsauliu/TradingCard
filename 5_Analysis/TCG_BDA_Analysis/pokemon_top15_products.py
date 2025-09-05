#!/usr/bin/env python3
"""
Top 15 Pokemon Products Analysis
Reads product IDs from top15products.csv and analyzes them
Outputs:
1. ASP weighted by Volume and Condition for each card
2. Trading Volumes for each card

Usage:
    python3 pokemon_top15_products.py [--scrape_date YYYY-MM-DD]
    
Example:
    python3 pokemon_top15_products.py --scrape_date 2025-09-04
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
import argparse
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def validate_scrape_date(client, scrape_date):
    """Validate that scrape_date exists in the database"""
    query = f"""
    SELECT COUNT(*) as row_count
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` bda
    INNER JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` meta
        ON CAST(bda.product_id AS STRING) = CAST(meta.product_productId AS STRING)
    WHERE meta.category_categoryId IN (3, 85)
        AND bda.scrape_date = '{scrape_date}'
    LIMIT 1
    """
    
    result = client.query(query).to_dataframe()
    return result['row_count'].iloc[0] > 0

def get_available_scrape_dates(client):
    """Get list of available scrape_dates"""
    query = """
    SELECT DISTINCT scrape_date
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda`
    ORDER BY scrape_date DESC
    LIMIT 10
    """
    
    result = client.query(query).to_dataframe()
    return result['scrape_date'].tolist()

def read_product_ids_from_csv(csv_path='top15products.csv'):
    """Read product IDs from CSV file containing TCGPlayer URLs, preserving order"""
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    product_ids = []  # Will maintain CSV order
    with open(csv_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                # Extract product ID from URL: /product/453466/
                match = re.search(r'/product/(\d+)/', line)
                if match:
                    product_ids.append(int(match.group(1)))
    
    if not product_ids:
        raise ValueError(f"No valid product IDs found in {csv_path}")
    
    print(f"Read {len(product_ids)} product IDs from {csv_path} (order preserved)")
    return product_ids

def add_formatted_section(ws, df, start_row, title, is_price=False):
    """Add formatted data section to worksheet"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Merge title across columns
    if not df.empty:
        ws.merge_cells(start_row=start_row, start_column=1, 
                      end_row=start_row, end_column=len(df.columns))
    
    # Headers
    header_row = start_row + 2
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    data_start_row = header_row + 1
    for row_idx, row in enumerate(df.itertuples(index=False), data_start_row):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, (int, float)) and col_idx > 1:
                if is_price:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
            
            # Alternate row colors
            if (row_idx - data_start_row) % 2 == 1:
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            # Alignment
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 10
        column_letter = get_column_letter(col_idx)
        
        for row in range(header_row, row_idx + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                # For product names, cap at 25 characters
                if col_idx == 1:
                    max_length = max(max_length, min(len(str(cell_value)) + 2, 30))
                else:
                    max_length = max(max_length, len(str(cell_value)) + 2)
        
        ws.column_dimensions[column_letter].width = min(max_length, 25)
    
    return row_idx + 3


def main():
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Top 15 Pokemon Products Analysis')
    parser.add_argument('--scrape_date', type=str, 
                       help='Scrape date to analyze (YYYY-MM-DD format)',
                       default='2025-09-04')
    
    args = parser.parse_args()
    scrape_date = args.scrape_date
    
    # Validate date format
    try:
        datetime.strptime(scrape_date, '%Y-%m-%d')
    except ValueError:
        print(f"Error: Invalid date format '{scrape_date}'. Please use YYYY-MM-DD format.")
        sys.exit(1)
    
    # Setup
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/5_Analysis/service-account.json'
    client = bigquery.Client()
    
    # Validate scrape_date exists
    print(f"Validating scrape_date={scrape_date}...")
    if not validate_scrape_date(client, scrape_date):
        print(f"Error: No data found for scrape_date={scrape_date}")
        print("\nAvailable scrape_dates:")
        available_dates = get_available_scrape_dates(client)
        for date in available_dates:
            print(f"  - {date}")
        sys.exit(1)
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    scrape_date_clean = scrape_date.replace('-', '')  # For filename
    
    # Read product IDs from CSV file
    print(f"Reading product IDs from top15products.csv...")
    try:
        product_ids = read_product_ids_from_csv('top15products.csv')
    except (FileNotFoundError, ValueError) as e:
        print(f"Error: {e}")
        sys.exit(1)
    
    product_ids_str = "', '".join([str(pid) for pid in product_ids])
    print(f"Processing {len(product_ids)} products: {product_ids}")
    
    # Get metadata for these specific products
    metadata_query = f"""
    SELECT DISTINCT
        bda.product_id,
        MAX(meta.product_name) as product_name,
        MAX(meta.product_cleanName) as product_clean_name,
        MAX(meta.group_name) as group_name,
        SUM(bda.quantity_sold) as lifecycle_quantity_sold,
        CASE 
            WHEN SUM(bda.quantity_sold) > 0 
            THEN SUM(bda.quantity_sold * bda.market_price) / SUM(bda.quantity_sold)
            ELSE 0 
        END as lifecycle_asp
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` bda
    INNER JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` meta
        ON CAST(bda.product_id AS STRING) = CAST(meta.product_productId AS STRING)
    WHERE bda.product_id IN ('{product_ids_str}')
        AND bda.bucket_start_date >= '2024-01-01'
        AND bda.scrape_date = '{scrape_date}'
    GROUP BY bda.product_id
    ORDER BY bda.product_id
    """
    
    print("Fetching product metadata...")
    top_products_df = client.query(metadata_query).to_dataframe()
    
    if top_products_df.empty:
        print(f"No data found for these products with scrape_date={scrape_date}")
        return
    
    # Convert product_id to int to match CSV IDs
    top_products_df['product_id'] = top_products_df['product_id'].astype(int)
    
    # Create order mapping from CSV
    order_mapping = {pid: idx for idx, pid in enumerate(product_ids)}
    
    # Sort dataframe to match CSV order
    top_products_df['csv_order'] = top_products_df['product_id'].map(order_mapping)
    top_products_df = top_products_df.sort_values('csv_order')
    top_products_df = top_products_df.drop('csv_order', axis=1)
    top_products_df = top_products_df.reset_index(drop=True)  # Reset index after sorting
    
    print(f"Found data for {len(top_products_df)} products (sorted to match CSV order)")
    
    # Get weekly data for these products
    weekly_data_query = f"""
    SELECT 
        bda.product_id,
        MAX(meta.product_cleanName) as product_name,
        DATE_TRUNC(bda.bucket_start_date, WEEK(MONDAY)) as week_start,
        -- Weighted ASP
        CASE 
            WHEN SUM(bda.quantity_sold) > 0 
            THEN SUM(bda.quantity_sold * bda.market_price) / SUM(bda.quantity_sold)
            ELSE 0 
        END as weighted_asp,
        -- Total volume
        SUM(bda.quantity_sold) as total_volume
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` bda
    LEFT JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` meta
        ON CAST(bda.product_id AS STRING) = CAST(meta.product_productId AS STRING)
    WHERE bda.product_id IN ('{product_ids_str}')
        AND bda.bucket_start_date >= '2024-01-01'
        AND bda.market_price > 0
        AND bda.quantity_sold IS NOT NULL
        AND bda.scrape_date = '{scrape_date}'  -- Specific scrape_date filter
    GROUP BY bda.product_id, DATE_TRUNC(bda.bucket_start_date, WEEK(MONDAY))
    ORDER BY bda.product_id, week_start
    """
    
    print("Fetching weekly data for top products...")
    weekly_df = client.query(weekly_data_query).to_dataframe()
    
    if weekly_df.empty:
        print("No weekly data found.")
        return
    
    # Convert product_id to int to match CSV IDs
    weekly_df['product_id'] = weekly_df['product_id'].astype(int)
    
    print(f"Retrieved {len(weekly_df)} weekly records")
    
    # Create order mapping from CSV for sorting
    order_mapping = {pid: idx for idx, pid in enumerate(product_ids)}
    
    # Create pivot tables
    # 1. ASP pivot - products as rows, weeks as columns
    asp_pivot = weekly_df.pivot_table(
        index=['product_id', 'product_name'],
        columns='week_start',
        values='weighted_asp',
        fill_value=0
    ).round(2)
    
    # Format column headers as dates
    asp_pivot.columns = [col.strftime('%Y-%m-%d') for col in asp_pivot.columns]
    asp_pivot = asp_pivot.reset_index()
    
    # Sort by CSV order
    asp_pivot['csv_order'] = asp_pivot['product_id'].map(order_mapping)
    asp_pivot = asp_pivot.sort_values('csv_order').drop('csv_order', axis=1)
    
    # Create display names for products (truncated for readability)
    asp_pivot['Product'] = asp_pivot.apply(
        lambda x: f"{x['product_name'][:40]}... ({x['product_id']})" 
        if len(str(x['product_name'])) > 40 
        else f"{x['product_name']} ({x['product_id']})", 
        axis=1
    )
    
    # Reorder columns
    asp_columns = ['Product'] + [col for col in asp_pivot.columns if col not in ['Product', 'product_id', 'product_name']]
    asp_display = asp_pivot[asp_columns].copy()
    
    # 2. Volume pivot - products as rows, weeks as columns  
    volume_pivot = weekly_df.pivot_table(
        index=['product_id', 'product_name'],
        columns='week_start',
        values='total_volume',
        fill_value=0
    ).astype(int)
    
    # Format column headers as dates
    volume_pivot.columns = [col.strftime('%Y-%m-%d') for col in volume_pivot.columns]
    volume_pivot = volume_pivot.reset_index()
    
    # Sort by CSV order
    volume_pivot['csv_order'] = volume_pivot['product_id'].map(order_mapping)
    volume_pivot = volume_pivot.sort_values('csv_order').drop('csv_order', axis=1)
    
    # Create display names
    volume_pivot['Product'] = volume_pivot.apply(
        lambda x: f"{x['product_name'][:40]}... ({x['product_id']})" 
        if len(str(x['product_name'])) > 40 
        else f"{x['product_name']} ({x['product_id']})", 
        axis=1
    )
    
    # Reorder columns
    volume_columns = ['Product'] + [col for col in volume_pivot.columns if col not in ['Product', 'product_id', 'product_name']]
    volume_display = volume_pivot[volume_columns].copy()
    
    # Create output directory
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create Excel file with scrape_date at the beginning of filename
    excel_file = f"{output_dir}/sd{scrape_date_clean}_{timestamp}_pokemon_top15_products.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 15 Pokemon Products"
    
    # Add scrape_date info at top
    info_cell = ws.cell(row=1, column=1, value=f"Data Source: scrape_date={scrape_date}")
    info_cell.font = Font(italic=True, size=10)
    
    # Add formatted sections
    current_row = 3
    
    # Section 1: Weighted ASP
    current_row = add_formatted_section(
        ws, asp_display, current_row,
        "1. Average Selling Price (ASP) - Weighted by Volume and Condition",
        is_price=True
    )
    
    # Section 2: Trading Volumes
    current_row = add_formatted_section(
        ws, volume_display, current_row,
        "2. Trading Volumes",
        is_price=False
    )
    
    # Add metadata sheet
    ws_meta = wb.create_sheet("Product Metadata")
    
    # Prepare metadata for display
    meta_display = top_products_df[['product_id', 'product_name', 'group_name', 
                                    'lifecycle_quantity_sold', 'lifecycle_asp']].copy()
    meta_display.columns = ['Product ID', 'Product Name', 'Set/Group', 
                            'Lifetime Sales', 'Lifetime ASP']
    meta_display['Rank'] = range(1, len(meta_display) + 1)
    meta_display = meta_display[['Rank', 'Product ID', 'Product Name', 'Set/Group', 
                                  'Lifetime Sales', 'Lifetime ASP']]
    
    add_formatted_section(ws_meta, meta_display, 1, "Top 15 Pokemon Products (EN + JP) - Metadata", is_price=False)
    
    # Format lifetime ASP column as currency
    for row in range(4, 4 + len(meta_display)):  # Starting from data row
        ws_meta.cell(row=row, column=6).number_format = '$#,##0.00'
    
    # Save Excel file
    wb.save(excel_file)
    
    print(f"\nExcel file saved: {excel_file}")
    print(f"Number of products: {len(top_products_df)}")
    print(f"Week range: {asp_display.columns[1]} to {asp_display.columns[-1]}")
    
    # Print products summary in CSV order
    print("\nProducts in CSV order:")
    print("-" * 80)
    for idx in range(len(top_products_df)):
        row = top_products_df.iloc[idx]
        product_name = row['product_clean_name'][:45] if len(row['product_clean_name']) > 45 else row['product_clean_name']
        print(f"{idx+1:2}. {product_name:45} | ID: {row['product_id']:6} | Sales: {row['lifecycle_quantity_sold']:7,} | ASP: ${row['lifecycle_asp']:6.2f}")

if __name__ == "__main__":
    main()