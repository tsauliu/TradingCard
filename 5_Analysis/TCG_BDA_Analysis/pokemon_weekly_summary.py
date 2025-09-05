#!/usr/bin/env python3
"""
Pokemon Weekly Summary Analysis
Outputs:
1. ASP weighted by Condition and Volume
2. Trading Volumes by Condition
3. Unique Pokemon Products with Sales
"""

import pandas as pd
from google.cloud import bigquery
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def add_formatted_section(ws, df, start_row, title, is_price=False):
    """Add formatted data section to worksheet"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title with formatting
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Merge title across columns
    if not df.empty:
        ws.merge_cells(start_row=start_row, start_column=1, 
                      end_row=start_row, end_column=len(df.columns))
    
    # Headers row
    header_row = start_row + 2
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Data rows
    data_start_row = header_row + 1
    for row_idx, row in enumerate(df.itertuples(index=False), data_start_row):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, (int, float)) and col_idx > 1:
                if is_price:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0'
            
            # Alternate row shading
            if (row_idx - data_start_row) % 2 == 1:
                cell.fill = PatternFill(start_color="E9EDF7", end_color="E9EDF7", fill_type="solid")
            
            # Left align condition column, right align numbers
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 12  # minimum width
        column_letter = get_column_letter(col_idx)
        
        # Check all cells in column
        for row in range(header_row, row_idx + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)) + 2)
        
        ws.column_dimensions[column_letter].width = min(max_length, 20)
    
    return row_idx + 3

def main():
    # Setup
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = '/home/caoliu/TradingCard/5_Analysis/service-account.json'
    client = bigquery.Client()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    print("Fetching Pokemon weekly data (scrape_date=2025-08-20)...")
    
    # 1. Weighted ASP by Condition and Volume
    asp_query = """
    WITH pokemon_weekly AS (
        SELECT 
            condition,
            DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
            SUM(quantity_sold * market_price) as total_value,
            SUM(quantity_sold) as total_quantity
        FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` bda
        INNER JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` meta
            ON CAST(bda.product_id AS STRING) = CAST(meta.product_productId AS STRING)
        WHERE meta.category_categoryId IN (3, 85)  -- Pokemon and Pokemon Japan categories
            AND bda.market_price > 0 
            AND bda.quantity_sold > 0
            AND bda.condition IS NOT NULL 
            AND bda.condition != ''
            AND bda.bucket_start_date >= '2024-01-01'
            AND bda.scrape_date = '2025-08-20'  -- Specific scrape_date filter
        GROUP BY condition, DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    )
    SELECT 
        condition,
        week_start,
        CASE 
            WHEN total_quantity > 0 THEN total_value / total_quantity 
            ELSE 0 
        END as weighted_asp
    FROM pokemon_weekly
    ORDER BY condition, week_start
    """
    
    # 2. Trading Volumes by Condition
    volume_query = """
    SELECT 
        condition,
        DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
        SUM(quantity_sold) as trading_volume
    FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` bda
    INNER JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` meta
        ON CAST(bda.product_id AS STRING) = CAST(meta.product_productId AS STRING)
    WHERE meta.category_categoryId IN (3, 85)  -- Pokemon and Pokemon Japan categories
        AND bda.quantity_sold IS NOT NULL
        AND bda.condition IS NOT NULL 
        AND bda.condition != ''
        AND bda.bucket_start_date >= '2024-01-01'
        AND bda.scrape_date = '2025-08-20'  -- Specific scrape_date filter
    GROUP BY condition, DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    ORDER BY condition, week_start
    """
    
    # 3. Unique Pokemon Products with Sales
    unique_query = """
    WITH weekly_unique AS (
        SELECT 
            DATE_TRUNC(bucket_start_date, WEEK(MONDAY)) as week_start,
            COUNT(DISTINCT bda.product_id) as unique_products
        FROM `rising-environs-456314-a3.tcg_data.tcg_prices_bda` bda
        INNER JOIN `rising-environs-456314-a3.tcg_data.tcg_metadata` meta
            ON CAST(bda.product_id AS STRING) = CAST(meta.product_productId AS STRING)
        WHERE meta.category_categoryId IN (3, 85)  -- Pokemon and Pokemon Japan categories
            AND bda.quantity_sold > 0
            AND bda.bucket_start_date >= '2024-01-01'
            AND bda.scrape_date = '2025-08-20'  -- Specific scrape_date filter
        GROUP BY DATE_TRUNC(bucket_start_date, WEEK(MONDAY))
    )
    SELECT 
        'Total Pokemon Products (EN + JP)' as metric,
        week_start,
        unique_products
    FROM weekly_unique
    ORDER BY week_start
    """
    
    # Execute queries
    print("Executing ASP query...")
    asp_df = client.query(asp_query).to_dataframe()
    
    print("Executing volume query...")
    volume_df = client.query(volume_query).to_dataframe()
    
    print("Executing unique products query...")
    unique_df = client.query(unique_query).to_dataframe()
    
    print(f"ASP data: {len(asp_df)} rows")
    print(f"Volume data: {len(volume_df)} rows")
    print(f"Unique products data: {len(unique_df)} rows")
    
    if asp_df.empty or volume_df.empty or unique_df.empty:
        print("Warning: Some queries returned no data")
    
    # Pivot data - weeks as columns
    # ASP pivot
    asp_pivot = asp_df.pivot(
        index='condition', 
        columns='week_start', 
        values='weighted_asp'
    ).round(2)
    asp_pivot.columns = [col.strftime('%Y-%m-%d') for col in asp_pivot.columns]
    asp_pivot = asp_pivot.reset_index()
    asp_pivot.columns.name = None
    
    # Volume pivot
    volume_pivot = volume_df.pivot(
        index='condition', 
        columns='week_start', 
        values='trading_volume'
    ).fillna(0).astype(int)
    volume_pivot.columns = [col.strftime('%Y-%m-%d') for col in volume_pivot.columns]
    volume_pivot = volume_pivot.reset_index()
    volume_pivot.columns.name = None
    
    # Unique products pivot
    unique_pivot = unique_df.pivot(
        index='metric', 
        columns='week_start', 
        values='unique_products'
    ).fillna(0).astype(int)
    unique_pivot.columns = [col.strftime('%Y-%m-%d') for col in unique_pivot.columns]
    unique_pivot = unique_pivot.reset_index()
    unique_pivot.columns.name = None
    
    # Create output directory
    output_dir = "../output"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create Excel file
    excel_file = f"{output_dir}/{timestamp}_pokemon_weekly_summary.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pokemon Weekly Summary"
    
    # Add formatted sections
    current_row = 1
    
    # Section 1: Weighted ASP
    current_row = add_formatted_section(
        ws, asp_pivot, current_row, 
        "1. Average Selling Price (ASP) - Weighted by Condition and Volume", 
        is_price=True
    )
    
    # Section 2: Trading Volumes
    current_row = add_formatted_section(
        ws, volume_pivot, current_row, 
        "2. Trading Volumes by Condition", 
        is_price=False
    )
    
    # Section 3: Unique Products
    current_row = add_formatted_section(
        ws, unique_pivot, current_row, 
        "3. Unique Pokemon Products with Sales", 
        is_price=False
    )
    
    # Save file
    wb.save(excel_file)
    
    print(f"\nExcel file saved: {excel_file}")
    print(f"ASP conditions: {sorted(asp_pivot['condition'].tolist())}")
    print(f"Volume conditions: {sorted(volume_pivot['condition'].tolist())}")
    print(f"Week range: {asp_pivot.columns[1]} to {asp_pivot.columns[-1]}")
    
    # Summary statistics
    print("\nSummary Statistics:")
    if not asp_df.empty:
        overall_asp = asp_df.groupby('week_start')['weighted_asp'].mean().mean()
        print(f"Overall average ASP: ${overall_asp:.2f}")
    
    if not volume_df.empty:
        total_volume = volume_df['trading_volume'].sum()
        print(f"Total trading volume: {total_volume:,}")
    
    if not unique_df.empty:
        avg_unique = unique_df['unique_products'].mean()
        print(f"Average unique products per week: {avg_unique:.0f}")

if __name__ == "__main__":
    main()