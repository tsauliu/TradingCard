#!/usr/bin/env python3
"""
Excel Pivot Table Utilities for eBay Search Data
Shared functions for converting eBay API data to pivot table format
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

logger = logging.getLogger(__name__)


def timestamp_to_week_string(timestamp_ms: int) -> str:
    """
    Convert millisecond timestamp to week string format
    
    Args:
        timestamp_ms: Timestamp in milliseconds
    
    Returns:
        Week string in format 'YYYY-MM-DD' (Monday of the week)
    """
    dt = datetime.fromtimestamp(timestamp_ms / 1000)
    # Get Monday of the week
    monday = dt - timedelta(days=dt.weekday())
    return monday.strftime('%Y-%m-%d')


def extract_series_data(api_response: Dict[str, Any]) -> Tuple[Dict[str, List], Dict[str, List]]:
    """
    Extract price and quantity series from API response
    
    Args:
        api_response: Raw API response dictionary
    
    Returns:
        Tuple of (price_data, quantity_data) where each is {week: [values]}
    """
    price_data = {}
    quantity_data = {}
    
    # Find MetricsTrendsModule
    metrics_module = api_response.get('MetricsTrendsModule')
    
    if not metrics_module:
        # Search in values if not found by key
        for item in api_response.values():
            if isinstance(item, dict) and item.get('_type') == 'MetricsTrendsModule':
                metrics_module = item
                break
    
    if metrics_module:
        series = metrics_module.get('series', [])
        
        for serie in series:
            if serie.get('id') == 'averageSold':
                # Extract average prices
                for data_point in serie.get('data', []):
                    if len(data_point) >= 2 and data_point[1] is not None:
                        week = timestamp_to_week_string(data_point[0])
                        price_data[week] = data_point[1]
            
            elif serie.get('id') == 'quantity':
                # Extract quantities
                for data_point in serie.get('data', []):
                    if len(data_point) >= 2 and data_point[1] is not None:
                        week = timestamp_to_week_string(data_point[0])
                        quantity_data[week] = data_point[1]
    
    return price_data, quantity_data


def create_pivot_dataframes(
    search_results: List[Dict[str, Any]], 
    fill_missing: bool = True
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Create pivot DataFrames from search results
    
    Args:
        search_results: List of search results with 'keywords' and 'raw_data'/'data'
        fill_missing: Whether to fill missing values with NaN or 0
    
    Returns:
        Tuple of (price_df, quantity_df) in pivot format
    """
    all_price_data = {}
    all_quantity_data = {}
    all_weeks = set()
    
    # Process each search result
    for result in search_results:
        keywords = result.get('keywords', 'Unknown')
        
        # Handle both single search and batch search formats
        if 'raw_data' in result:
            api_data = result['raw_data']
        else:
            api_data = result.get('data', result)
        
        price_data, quantity_data = extract_series_data(api_data)
        
        all_price_data[keywords] = price_data
        all_quantity_data[keywords] = quantity_data
        all_weeks.update(price_data.keys())
        all_weeks.update(quantity_data.keys())
    
    # Sort weeks chronologically
    sorted_weeks = sorted(list(all_weeks))
    
    # Create DataFrames
    price_rows = []
    quantity_rows = []
    
    for keywords in all_price_data.keys():
        price_row = {'Keywords': keywords}
        quantity_row = {'Keywords': keywords}
        
        for week in sorted_weeks:
            price_row[week] = all_price_data[keywords].get(week, np.nan if fill_missing else 0)
            quantity_row[week] = all_quantity_data[keywords].get(week, np.nan if fill_missing else 0)
        
        price_rows.append(price_row)
        quantity_rows.append(quantity_row)
    
    price_df = pd.DataFrame(price_rows)
    quantity_df = pd.DataFrame(quantity_rows)
    
    # Set Keywords as index
    if not price_df.empty:
        price_df.set_index('Keywords', inplace=True)
        quantity_df.set_index('Keywords', inplace=True)
    
    return price_df, quantity_df


def save_pivot_to_excel(
    price_df: pd.DataFrame,
    quantity_df: pd.DataFrame,
    output_file: str,
    metadata: Optional[Dict] = None,
    add_charts: bool = True,
    add_formatting: bool = True
):
    """
    Save pivot DataFrames to Excel with formatting
    
    Args:
        price_df: Price pivot DataFrame
        quantity_df: Quantity pivot DataFrame
        output_file: Output Excel file path
        metadata: Optional metadata to include
        add_charts: Whether to add trend charts
        add_formatting: Whether to add cell formatting
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write DataFrames
        price_df.to_excel(writer, sheet_name='Prices', freeze_panes=(1, 1))
        quantity_df.to_excel(writer, sheet_name='Quantities', freeze_panes=(1, 1))
        
        # Add statistics sheet
        stats_data = []
        for keywords in price_df.index:
            stats_row = {
                'Keywords': keywords,
                'Avg Price': price_df.loc[keywords].mean(),
                'Min Price': price_df.loc[keywords].min(),
                'Max Price': price_df.loc[keywords].max(),
                'Price Std Dev': price_df.loc[keywords].std(),
                'Total Quantity': quantity_df.loc[keywords].sum(),
                'Avg Weekly Quantity': quantity_df.loc[keywords].mean(),
                'Data Points': price_df.loc[keywords].notna().sum()
            }
            stats_data.append(stats_row)
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)
        
        # Add metadata sheet if provided
        if metadata:
            meta_df = pd.DataFrame(list(metadata.items()), columns=['Parameter', 'Value'])
            meta_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Get workbook for formatting
        workbook = writer.book
        
        if add_formatting:
            # Format price sheet
            price_sheet = workbook['Prices']
            format_pivot_sheet(price_sheet, data_format='currency')
            
            # Format quantity sheet
            quantity_sheet = workbook['Quantities']
            format_pivot_sheet(quantity_sheet, data_format='number')
            
            # Format statistics sheet
            stats_sheet = workbook['Statistics']
            format_stats_sheet(stats_sheet)
        
        if add_charts and len(price_df.columns) > 1:
            # Add price trend chart
            add_trend_chart(
                workbook, 
                'Prices', 
                'Price Trends',
                'Week', 
                'Average Price ($)',
                'Price_Chart'
            )
            
            # Add quantity trend chart
            add_trend_chart(
                workbook, 
                'Quantities', 
                'Sales Volume Trends',
                'Week', 
                'Quantity Sold',
                'Quantity_Chart'
            )
    
    logger.info(f"Excel pivot table saved to: {output_file}")


def format_pivot_sheet(sheet, data_format='number'):
    """
    Apply formatting to pivot table sheet
    
    Args:
        sheet: Openpyxl worksheet
        data_format: 'currency', 'number', or 'percent'
    """
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Apply header formatting
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Keywords column formatting
    keywords_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for row in sheet.iter_rows(min_row=2, max_col=1):
        for cell in row:
            cell.fill = keywords_fill
            cell.font = Font(bold=True)
    
    # Data cell formatting
    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if cell.value is not None and not isinstance(cell.value, str):
                if data_format == 'currency':
                    cell.number_format = '$#,##0.00'
                elif data_format == 'number':
                    cell.number_format = '#,##0'
                elif data_format == 'percent':
                    cell.number_format = '0.00%'
                
                # Conditional coloring based on value
                if isinstance(cell.value, (int, float)):
                    if data_format == 'currency' and cell.value > 50:
                        cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                    elif data_format == 'currency' and cell.value < 10:
                        cell.fill = PatternFill(start_color="E5FFE5", end_color="E5FFE5", fill_type="solid")
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def format_stats_sheet(sheet):
    """Format the statistics sheet"""
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format numeric columns
    for row in sheet.iter_rows(min_row=2):
        for idx, cell in enumerate(row):
            if idx > 0 and cell.value is not None and not isinstance(cell.value, str):
                if 'Price' in sheet.cell(1, idx + 1).value:
                    cell.number_format = '$#,##0.00'
                elif 'Quantity' in sheet.cell(1, idx + 1).value:
                    cell.number_format = '#,##0'
                elif 'Points' in sheet.cell(1, idx + 1).value:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'
    
    # Auto-adjust columns
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        sheet.column_dimensions[column_letter].width = adjusted_width


def add_trend_chart(workbook, sheet_name, chart_title, x_label, y_label, chart_sheet_name):
    """
    Add a trend chart to the workbook
    
    Args:
        workbook: Openpyxl workbook
        sheet_name: Source data sheet name
        chart_title: Title for the chart
        x_label: X-axis label
        y_label: Y-axis label
        chart_sheet_name: Name for the new chart sheet
    """
    ws = workbook[sheet_name]
    
    # Create chart
    chart = LineChart()
    chart.title = chart_title
    chart.style = 10
    chart.y_axis.title = y_label
    chart.x_axis.title = x_label
    chart.height = 15
    chart.width = 25
    
    # Add data series for each keyword
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Categories (dates)
    dates = Reference(ws, min_col=2, min_row=1, max_col=max_col)
    
    # Add each row as a series
    for row in range(2, max_row + 1):
        values = Reference(ws, min_col=2, min_row=row, max_col=max_col)
        chart.add_data(values, titles_from_data=False)
        # Set the series title
        if chart.series and len(chart.series) >= row - 1:
            chart.series[row - 2].title = str(ws.cell(row, 1).value)
    
    chart.set_categories(dates)
    
    # Create new sheet for chart
    chart_ws = workbook.create_sheet(chart_sheet_name)
    chart_ws.add_chart(chart, "A1")


def create_summary_pivot(
    search_results: List[Dict[str, Any]],
    output_file: str,
    time_period: str = 'weekly',
    add_statistics: bool = True,
    add_charts: bool = True
):
    """
    Create a comprehensive pivot table Excel file from search results
    
    Args:
        search_results: List of search results
        output_file: Output Excel file path
        time_period: 'weekly', 'monthly', or 'quarterly'
        add_statistics: Whether to add statistics sheet
        add_charts: Whether to add trend charts
    """
    # Create pivot DataFrames
    price_df, quantity_df = create_pivot_dataframes(search_results)
    
    # Aggregate by time period if needed
    if time_period == 'monthly' and not price_df.empty:
        # Convert column names to months
        price_df = aggregate_to_monthly(price_df)
        quantity_df = aggregate_to_monthly(quantity_df)
    elif time_period == 'quarterly' and not price_df.empty:
        # Convert to quarters
        price_df = aggregate_to_quarterly(price_df)
        quantity_df = aggregate_to_quarterly(quantity_df)
    
    # Prepare metadata
    metadata = {
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Total Keywords': len(search_results),
        'Time Period': time_period,
        'Data Points': len(price_df.columns) if not price_df.empty else 0
    }
    
    # Save to Excel
    save_pivot_to_excel(
        price_df,
        quantity_df,
        output_file,
        metadata=metadata,
        add_charts=add_charts,
        add_formatting=True
    )


def aggregate_to_monthly(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate weekly data to monthly"""
    if df.empty:
        return df
    
    # Group columns by month
    monthly_data = {}
    
    for col in df.columns:
        try:
            date = pd.to_datetime(col)
            month_key = date.strftime('%Y-%m')
            if month_key not in monthly_data:
                monthly_data[month_key] = []
            monthly_data[month_key].append(col)
        except:
            continue
    
    # Create new DataFrame with monthly averages
    monthly_df = pd.DataFrame(index=df.index)
    
    for month, week_cols in sorted(monthly_data.items()):
        monthly_df[month] = df[week_cols].mean(axis=1)
    
    return monthly_df


def aggregate_to_quarterly(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate weekly data to quarterly"""
    if df.empty:
        return df
    
    # Group columns by quarter
    quarterly_data = {}
    
    for col in df.columns:
        try:
            date = pd.to_datetime(col)
            quarter_key = f"{date.year}-Q{(date.month-1)//3 + 1}"
            if quarter_key not in quarterly_data:
                quarterly_data[quarter_key] = []
            quarterly_data[quarter_key].append(col)
        except:
            continue
    
    # Create new DataFrame with quarterly averages
    quarterly_df = pd.DataFrame(index=df.index)
    
    for quarter, week_cols in sorted(quarterly_data.items()):
        quarterly_df[quarter] = df[week_cols].mean(axis=1)
    
    return quarterly_df